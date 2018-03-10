#! python3
# blankRowInserter.py - insert  blank rows into the spreadsheet

import openpyxl, sys

try:
    rowNum = int(sys.argv[1])
    insertNum = int(sys.argv[2])
    file = sys.argv[3]
except: 
    rowNum = 3
    insertNum = 2
    file = 'table_6.xlsx'

wb = openpyxl.load_workbook(file)
sheet = wb.active

newwb = openpyxl.Workbook()
newsheet = newwb.active

print('insert %s blank rows after row %s in %s' % (str(insertNum),str(rowNum),file))

for beforeRow in range(1, rowNum):
    for columnObj in range(1, sheet.max_column+1):
        newsheet.cell(row=beforeRow, column=columnObj).value = sheet.cell(row=beforeRow, column=columnObj).value

for afterRow in range(rowNum, sheet.max_row+1):
    for newcolumnObj in range(1, sheet.max_column+1):
        newsheet.cell(row=afterRow+insertNum, column=newcolumnObj).value = sheet.cell(row=afterRow, column=newcolumnObj).value

newwb.save('insertTable.xlsx')
print('done')

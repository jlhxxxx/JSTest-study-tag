#! python3
# multiplicationTable.py - takes a number N from the command line and creates an N×N multiplication table in an Excel spreadsheet

import openpyxl, sys
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active
boldFont = Font(bold=True)
try:
    num = int(sys.argv[1])
except:
    num = 6   

print('creating a %s * %s table...' % (str(num), str(num)))

# write Row 1 and column A
for count in range(2, num + 2):
    sheet['A'+str(count)] = count - 1
    sheet['A'+str(count)].font = boldFont
    sheet.cell(row=1, column=count).value = count - 1
    sheet.cell(row=1, column=count).font = boldFont

# write inside table
for rowObj in range(1, num + 1):
    for columnObj in range(1, num + 1):
        sheet.cell(row=rowObj+1, column=columnObj+1).value = sheet.cell(row=1, column=columnObj+1).value * sheet.cell(row=rowObj+1, column=1).value

wb.save('table_%s.xlsx' % str(num))
print('done')
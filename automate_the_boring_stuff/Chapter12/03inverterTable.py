#! python3
# inverterTable.py - change sheetData[x][y] to sheetData[y][x]

import openpyxl,sys

try:
    file = sys.argv[1]
except: 
    file = 'insertTable.xlsx'

wb = openpyxl.load_workbook(file)
sheet = wb.active

newwb = openpyxl.Workbook()
newsheet = newwb.active
sheetData = []

print('Invertering %s' % file)

for rowObj in range(1, sheet.max_row+1):
    sheetData.append([])
    for columnObj in range(1, sheet.max_column+1):
        sheetData[rowObj-1].insert(columnObj-1, sheet.cell(row=rowObj, column=columnObj).value) 


for rowObj1 in range(1, sheet.max_column+1):
    for columnObj1 in range(1, sheet.max_row+1):
        newsheet.cell(row=rowObj1, column=columnObj1).value = sheetData[columnObj1-1][rowObj1-1]

newwb.save('invertered.xlsx')
print('done')
